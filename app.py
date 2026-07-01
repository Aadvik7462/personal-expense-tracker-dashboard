from flask import Flask, render_template, request, redirect, Response, session, flash
import sqlite3
import os
import json
import csv
from io import StringIO, BytesIO
from functools import wraps
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.secret_key = "expense_tracker_secret_key"

DB_PATH = "database/expense.db"


def init_db():
    os.makedirs("database", exist_ok=True)

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            title TEXT NOT NULL,
            amount REAL NOT NULL,
            category TEXT NOT NULL,
            type TEXT NOT NULL,
            date TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS budgets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            month TEXT NOT NULL,
            budget REAL NOT NULL,
            UNIQUE(user_id, month),
            FOREIGN KEY(user_id) REFERENCES users(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS recurring_transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            title TEXT NOT NULL,
            amount REAL NOT NULL,
            category TEXT NOT NULL,
            type TEXT NOT NULL,
            frequency TEXT NOT NULL,
            start_date TEXT NOT NULL,
            is_active INTEGER DEFAULT 1,
            FOREIGN KEY(user_id) REFERENCES users(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS bill_reminders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            title TEXT NOT NULL,
            amount REAL NOT NULL,
            due_date TEXT NOT NULL,
            status TEXT DEFAULT 'Pending',
            FOREIGN KEY(user_id) REFERENCES users(id)
        )
    """)

    try:
        cursor.execute("ALTER TABLE transactions ADD COLUMN user_id INTEGER")
    except sqlite3.OperationalError:
        pass

    conn.commit()
    conn.close()


def get_connection():
    return sqlite3.connect(DB_PATH)


def login_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            flash("Please login first.")
            return redirect("/login")
        return func(*args, **kwargs)
    return wrapper


def get_summary():
    user_id = session.get("user_id")
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT SUM(amount) FROM transactions WHERE type='Income' AND user_id=?", (user_id,))
    income = cursor.fetchone()[0] or 0

    cursor.execute("SELECT SUM(amount) FROM transactions WHERE type='Expense' AND user_id=?", (user_id,))
    expense = cursor.fetchone()[0] or 0

    cursor.execute("SELECT COUNT(*) FROM transactions WHERE user_id=?", (user_id,))
    total_transactions = cursor.fetchone()[0] or 0

    conn.close()
    return income, expense, income - expense, total_transactions


def get_dashboard_stats():
    user_id = session.get("user_id")
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT MAX(amount) FROM transactions WHERE type='Income' AND user_id=?", (user_id,))
    highest_income = cursor.fetchone()[0] or 0

    cursor.execute("SELECT MAX(amount) FROM transactions WHERE type='Expense' AND user_id=?", (user_id,))
    highest_expense = cursor.fetchone()[0] or 0

    cursor.execute("SELECT AVG(amount) FROM transactions WHERE type='Expense' AND user_id=?", (user_id,))
    average_expense = cursor.fetchone()[0] or 0

    cursor.execute("""
        SELECT category, COUNT(*)
        FROM transactions
        WHERE user_id=?
        GROUP BY category
        ORDER BY COUNT(*) DESC
        LIMIT 1
    """, (user_id,))

    result = cursor.fetchone()
    most_used_category = result[0] if result else "N/A"

    conn.close()
    return highest_income, highest_expense, round(average_expense, 2), most_used_category


def get_budget_data():
    user_id = session.get("user_id")
    selected_month = request.args.get("budget_month", "")

    if not selected_month:
        selected_month = datetime.now().strftime("%Y-%m")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT budget
        FROM budgets
        WHERE user_id=? AND month=?
    """, (user_id, selected_month))

    result = cursor.fetchone()
    monthly_budget = result[0] if result else 0

    cursor.execute("""
        SELECT SUM(amount)
        FROM transactions
        WHERE user_id=?
        AND type='Expense'
        AND substr(date, 1, 7)=?
    """, (user_id, selected_month))

    spent_amount = cursor.fetchone()[0] or 0
    remaining_budget = monthly_budget - spent_amount

    if monthly_budget > 0:
        used_percentage = round((spent_amount / monthly_budget) * 100, 2)
    else:
        used_percentage = 0

    if monthly_budget == 0:
        budget_status = "No Budget Set"
    elif spent_amount > monthly_budget:
        budget_status = "Budget Exceeded"
    else:
        budget_status = "Within Budget"

    conn.close()

    return {
        "selected_month": selected_month,
        "monthly_budget": monthly_budget,
        "spent_amount": spent_amount,
        "remaining_budget": remaining_budget,
        "used_percentage": used_percentage,
        "budget_status": budget_status
    }


def get_monthly_summary():
    user_id = session.get("user_id")
    current_month = datetime.now().strftime("%Y-%m")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT SUM(amount)
        FROM transactions
        WHERE user_id=?
        AND type='Income'
        AND substr(date, 1, 7)=?
    """, (user_id, current_month))
    monthly_income = cursor.fetchone()[0] or 0

    cursor.execute("""
        SELECT SUM(amount)
        FROM transactions
        WHERE user_id=?
        AND type='Expense'
        AND substr(date, 1, 7)=?
    """, (user_id, current_month))
    monthly_expense = cursor.fetchone()[0] or 0

    monthly_savings = monthly_income - monthly_expense

    conn.close()
    return monthly_income, monthly_expense, monthly_savings


def get_top_expenses():
    user_id = session.get("user_id")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT title, amount, category, date
        FROM transactions
        WHERE user_id=?
        AND type='Expense'
        ORDER BY amount DESC
        LIMIT 5
    """, (user_id,))

    top_expenses = cursor.fetchall()

    conn.close()
    return top_expenses


def get_category_percentages():
    user_id = session.get("user_id")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT SUM(amount)
        FROM transactions
        WHERE user_id=?
        AND type='Expense'
    """, (user_id,))
    total_expense = cursor.fetchone()[0] or 0

    cursor.execute("""
        SELECT category, SUM(amount)
        FROM transactions
        WHERE user_id=?
        AND type='Expense'
        GROUP BY category
        ORDER BY SUM(amount) DESC
    """, (user_id,))

    category_data = cursor.fetchall()
    conn.close()

    percentages = []

    for category, amount in category_data:
        if total_expense > 0:
            percent = round((amount / total_expense) * 100, 2)
        else:
            percent = 0

        percentages.append({
            "category": category,
            "amount": amount,
            "percent": percent
        })

    return percentages


def get_financial_insights():
    insights = []

    income, expense, balance, total_transactions = get_summary()
    budget_data = get_budget_data()
    monthly_income, monthly_expense, monthly_savings = get_monthly_summary()
    top_expenses = get_top_expenses()
    category_percentages = get_category_percentages()

    if total_transactions == 0:
        insights.append("💡 Start by adding your first income or expense transaction.")
        return insights

    if budget_data["budget_status"] == "Budget Exceeded":
        exceeded_amount = abs(budget_data["remaining_budget"])
        insights.append(f"⚠️ You exceeded your monthly budget by ₹{exceeded_amount}.")

    elif budget_data["budget_status"] == "Within Budget":
        insights.append(f"✅ You are within your budget. Remaining budget is ₹{budget_data['remaining_budget']}.")

    else:
        insights.append("💡 Set a monthly budget to track your spending better.")

    if monthly_savings > 0:
        insights.append(f"💰 You saved ₹{monthly_savings} this month.")
    elif monthly_savings < 0:
        insights.append(f"⚠️ Your expenses are ₹{abs(monthly_savings)} more than your income this month.")
    else:
        insights.append("💡 Your monthly income and expenses are currently balanced.")

    if top_expenses:
        biggest = top_expenses[0]
        insights.append(f"📌 Your biggest expense is {biggest[0]} of ₹{biggest[1]}.")

    if category_percentages:
        top_category = category_percentages[0]
        insights.append(
            f"📊 {top_category['category']} is your highest spending category at {top_category['percent']}%."
        )

    if balance < 0:
        insights.append("⚠️ Your overall balance is negative. Try reducing expenses.")
    elif balance > 0:
        insights.append("✅ Your overall balance is positive. Good financial control.")

    return insights


def get_recurring_transactions():
    user_id = session.get("user_id")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, title, amount, category, type, frequency, start_date, is_active
        FROM recurring_transactions
        WHERE user_id=?
        ORDER BY id DESC
    """, (user_id,))

    recurring_transactions = cursor.fetchall()

    conn.close()
    return recurring_transactions


def get_bill_reminders():
    user_id = session.get("user_id")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, title, amount, due_date, status
        FROM bill_reminders
        WHERE user_id=?
        ORDER BY due_date ASC
    """, (user_id,))

    rows = cursor.fetchall()
    conn.close()

    today = datetime.today().date()
    reminders = []

    for row in rows:
        due_date = datetime.strptime(row[3], "%Y-%m-%d").date()

        if row[4] == "Paid":
            badge = "Paid"
        elif due_date < today:
            badge = "Overdue"
        elif due_date == today:
            badge = "Due Today"
        else:
            badge = "Upcoming"

        reminders.append({
            "id": row[0],
            "title": row[1],
            "amount": row[2],
            "due_date": row[3],
            "status": row[4],
            "badge": badge
        })

    return reminders
def get_calendar_data():
    user_id = session.get("user_id")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT title, amount, category, type, date
        FROM transactions
        WHERE user_id=?
        ORDER BY date ASC
    """, (user_id,))

    transactions = cursor.fetchall()

    cursor.execute("""
        SELECT title, amount, due_date, status
        FROM bill_reminders
        WHERE user_id=?
        ORDER BY due_date ASC
    """, (user_id,))

    bills = cursor.fetchall()

    conn.close()

    calendar_items = {}

    for t in transactions:
        date = t[4]

        if date not in calendar_items:
            calendar_items[date] = []

        calendar_items[date].append({
            "title": t[0],
            "amount": t[1],
            "category": t[2],
            "type": t[3],
            "kind": "transaction"
        })

    for b in bills:
        date = b[2]

        if date not in calendar_items:
            calendar_items[date] = []

        calendar_items[date].append({
            "title": b[0],
            "amount": b[1],
            "category": "Bill Reminder",
            "type": b[3],
            "kind": "bill"
        })

    return calendar_items

def get_ai_predictions():
    user_id = session.get("user_id")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT substr(date, 1, 7), SUM(amount)
        FROM transactions
        WHERE user_id=?
        AND type='Expense'
        GROUP BY substr(date, 1, 7)
        ORDER BY substr(date, 1, 7)
    """, (user_id,))

    monthly_expenses = cursor.fetchall()

    cursor.execute("""
        SELECT category, SUM(amount)
        FROM transactions
        WHERE user_id=?
        AND type='Expense'
        GROUP BY category
        ORDER BY SUM(amount) DESC
        LIMIT 1
    """, (user_id,))

    top_category = cursor.fetchone()

    conn.close()

    predictions = []

    if not monthly_expenses:
        predictions.append("🤖 Add more expense data to generate spending predictions.")
        return predictions

    total_expense = sum(row[1] for row in monthly_expenses)
    avg_monthly_expense = round(total_expense / len(monthly_expenses), 2)

    predictions.append(f"🤖 Predicted average monthly expense is ₹{avg_monthly_expense}.")

    if len(monthly_expenses) >= 2:
        last_month_expense = monthly_expenses[-1][1]
        previous_month_expense = monthly_expenses[-2][1]

        if last_month_expense > previous_month_expense:
            increase = round(last_month_expense - previous_month_expense, 2)
            predictions.append(f"📈 Your expenses increased by ₹{increase} compared to the previous month.")

        elif last_month_expense < previous_month_expense:
            decrease = round(previous_month_expense - last_month_expense, 2)
            predictions.append(f"📉 Your expenses decreased by ₹{decrease} compared to the previous month.")

        else:
            predictions.append("📊 Your expenses remained stable compared to the previous month.")

    budget_data = get_budget_data()

    if budget_data["monthly_budget"] > 0:
        if avg_monthly_expense > budget_data["monthly_budget"]:
            predictions.append("⚠️ Your average monthly expense is higher than your budget.")
        else:
            predictions.append("✅ Your average monthly expense is within your budget.")

    if top_category:
        predictions.append(
            f"💡 Your highest spending area is {top_category[0]}. Try reducing this category by 10%."
        )

    income, expense, balance, total_transactions = get_summary()

    if balance < 0:
        predictions.append("🚨 Your balance is negative. Focus on reducing expenses immediately.")
    elif balance > 0:
        predictions.append("✅ Your balance is positive. You can allocate some amount toward savings.")

    return predictions

def get_chart_data():
    user_id = session.get("user_id")
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT category, SUM(amount)
        FROM transactions
        WHERE type='Expense' AND user_id=?
        GROUP BY category
        ORDER BY SUM(amount) DESC
    """, (user_id,))
    category_data = cursor.fetchall()

    cursor.execute("""
        SELECT substr(date, 1, 7), SUM(amount)
        FROM transactions
        WHERE type='Expense' AND user_id=?
        GROUP BY substr(date, 1, 7)
        ORDER BY substr(date, 1, 7)
    """, (user_id,))
    monthly_expense_data = cursor.fetchall()

    cursor.execute("""
        SELECT substr(date, 1, 7), SUM(amount)
        FROM transactions
        WHERE type='Income' AND user_id=?
        GROUP BY substr(date, 1, 7)
        ORDER BY substr(date, 1, 7)
    """, (user_id,))
    monthly_income_data = cursor.fetchall()

    cursor.execute("""
        SELECT substr(date, 1, 4), SUM(amount)
        FROM transactions
        WHERE type='Expense' AND user_id=?
        GROUP BY substr(date, 1, 4)
        ORDER BY substr(date, 1, 4)
    """, (user_id,))
    yearly_expense_data = cursor.fetchall()

    cursor.execute("""
        SELECT substr(date, 1, 4), SUM(amount)
        FROM transactions
        WHERE type='Income' AND user_id=?
        GROUP BY substr(date, 1, 4)
        ORDER BY substr(date, 1, 4)
    """, (user_id,))
    yearly_income_data = cursor.fetchall()

    income, expense, balance, total_transactions = get_summary()

    monthly_income_dict = dict(monthly_income_data)
    monthly_expense_dict = dict(monthly_expense_data)

    all_months = sorted(set(monthly_income_dict.keys()) | set(monthly_expense_dict.keys()))

    comparison_income = []
    comparison_expense = []
    savings_values = []

    for month in all_months:
        inc = monthly_income_dict.get(month, 0)
        exp = monthly_expense_dict.get(month, 0)

        comparison_income.append(inc)
        comparison_expense.append(exp)
        savings_values.append(inc - exp)

    yearly_income_dict = dict(yearly_income_data)
    yearly_expense_dict = dict(yearly_expense_data)

    all_years = sorted(set(yearly_income_dict.keys()) | set(yearly_expense_dict.keys()))

    yearly_income_values = []
    yearly_expense_values = []

    for year in all_years:
        yearly_income_values.append(yearly_income_dict.get(year, 0))
        yearly_expense_values.append(yearly_expense_dict.get(year, 0))

    conn.close()

    return {
        "income_expense_labels": json.dumps(["Income", "Expense"]),
        "income_expense_values": json.dumps([income, expense]),

        "category_labels": json.dumps([row[0] for row in category_data]),
        "category_values": json.dumps([row[1] for row in category_data]),

        "expense_month_labels": json.dumps([row[0] for row in monthly_expense_data]),
        "expense_month_values": json.dumps([row[1] for row in monthly_expense_data]),

        "income_month_labels": json.dumps([row[0] for row in monthly_income_data]),
        "income_month_values": json.dumps([row[1] for row in monthly_income_data]),

        "comparison_month_labels": json.dumps(all_months),
        "comparison_income_values": json.dumps(comparison_income),
        "comparison_expense_values": json.dumps(comparison_expense),

        "savings_month_labels": json.dumps(all_months),
        "savings_values": json.dumps(savings_values),

        "yearly_labels": json.dumps(all_years),
        "yearly_income_values": json.dumps(yearly_income_values),
        "yearly_expense_values": json.dumps(yearly_expense_values)
    }


@app.route("/register", methods=["GET", "POST"])
def register():
    init_db()

    if request.method == "POST":
        name = request.form["name"]
        email = request.form["email"]
        password = request.form["password"]
        confirm_password = request.form["confirm_password"]

        if password != confirm_password:
            flash("Passwords do not match.")
            return redirect("/register")

        hashed_password = generate_password_hash(password)

        conn = get_connection()
        cursor = conn.cursor()

        try:
            cursor.execute(
                "INSERT INTO users (name, email, password) VALUES (?, ?, ?)",
                (name, email, hashed_password)
            )
            conn.commit()
            conn.close()
            flash("Registration successful. Please login.")
            return redirect("/login")
        except sqlite3.IntegrityError:
            conn.close()
            flash("Email already registered.")
            return redirect("/register")

    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    init_db()

    if request.method == "POST":
        email = request.form["email"]
        password = request.form["password"]

        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE email=?", (email,))
        user = cursor.fetchone()
        conn.close()

        if user and check_password_hash(user[3], password):
            session["user_id"] = user[0]
            session["user_name"] = user[1]
            flash("Login successful.")
            return redirect("/")

        flash("Invalid email or password.")
        return redirect("/login")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("Logged out successfully.")
    return redirect("/login")


@app.route("/set-budget", methods=["POST"])
@login_required
def set_budget():
    user_id = session.get("user_id")
    month = request.form["month"]
    budget = request.form["budget"]

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO budgets (user_id, month, budget)
        VALUES (?, ?, ?)
        ON CONFLICT(user_id, month)
        DO UPDATE SET budget=excluded.budget
    """, (user_id, month, budget))

    conn.commit()
    conn.close()

    flash("Budget saved successfully.")
    return redirect(f"/?budget_month={month}")


@app.route("/add-recurring", methods=["POST"])
@login_required
def add_recurring():
    user_id = session.get("user_id")

    title = request.form["title"]
    amount = request.form["amount"]
    category = request.form["category"]
    trans_type = request.form["type"]
    frequency = request.form["frequency"]
    start_date = request.form["start_date"]

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO recurring_transactions
        (user_id, title, amount, category, type, frequency, start_date)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (user_id, title, amount, category, trans_type, frequency, start_date))

    conn.commit()
    conn.close()

    flash("Recurring transaction added successfully.")
    return redirect("/")


@app.route("/toggle-recurring/<int:id>")
@login_required
def toggle_recurring(id):
    user_id = session.get("user_id")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT is_active
        FROM recurring_transactions
        WHERE id=? AND user_id=?
    """, (id, user_id))

    result = cursor.fetchone()

    if result:
        new_status = 0 if result[0] == 1 else 1

        cursor.execute("""
            UPDATE recurring_transactions
            SET is_active=?
            WHERE id=? AND user_id=?
        """, (new_status, id, user_id))

        conn.commit()
        flash("Recurring transaction status updated.")

    conn.close()
    return redirect("/")


@app.route("/generate-recurring/<int:id>")
@login_required
def generate_recurring(id):
    user_id = session.get("user_id")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT title, amount, category, type
        FROM recurring_transactions
        WHERE id=? AND user_id=? AND is_active=1
    """, (id, user_id))

    recurring = cursor.fetchone()

    if recurring:
        today = datetime.now().strftime("%Y-%m-%d")

        cursor.execute("""
            INSERT INTO transactions
            (user_id, title, amount, category, type, date)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (user_id, recurring[0], recurring[1], recurring[2], recurring[3], today))

        conn.commit()
        flash("Recurring transaction generated successfully.")
    else:
        flash("Recurring transaction not found or inactive.")

    conn.close()
    return redirect("/")


@app.route("/add-bill", methods=["POST"])
@login_required
def add_bill():
    user_id = session.get("user_id")

    title = request.form["title"]
    amount = request.form["amount"]
    due_date = request.form["due_date"]

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO bill_reminders (user_id, title, amount, due_date)
        VALUES (?, ?, ?, ?)
    """, (user_id, title, amount, due_date))

    conn.commit()
    conn.close()

    flash("Bill reminder added successfully.")
    return redirect("/")


@app.route("/bill-paid/<int:id>")
@login_required
def bill_paid(id):
    user_id = session.get("user_id")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE bill_reminders
        SET status='Paid'
        WHERE id=? AND user_id=?
    """, (id, user_id))

    conn.commit()
    conn.close()

    flash("Bill marked as paid.")
    return redirect("/")


@app.route("/delete-bill/<int:id>")
@login_required
def delete_bill(id):
    user_id = session.get("user_id")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        DELETE FROM bill_reminders
        WHERE id=? AND user_id=?
    """, (id, user_id))

    conn.commit()
    conn.close()

    flash("Bill reminder deleted.")
    return redirect("/")


@app.route("/")
@login_required
def dashboard():
    init_db()
    user_id = session.get("user_id")

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT * FROM transactions WHERE user_id=? ORDER BY id DESC LIMIT 5",
        (user_id,)
    )
    recent_transactions = cursor.fetchall()
    conn.close()

    income, expense, balance, total_transactions = get_summary()
    highest_income, highest_expense, average_expense, most_used_category = get_dashboard_stats()
    budget_data = get_budget_data()
    monthly_income, monthly_expense, monthly_savings = get_monthly_summary()
    top_expenses = get_top_expenses()
    category_percentages = get_category_percentages()
    recurring_transactions = get_recurring_transactions()
    bill_reminders = get_bill_reminders()
    financial_insights = get_financial_insights()
    ai_predictions = get_ai_predictions()

    return render_template(
        "dashboard.html",
        income=income,
        expense=expense,
        balance=balance,
        total_transactions=total_transactions,
        recent_transactions=recent_transactions,
        user_name=session.get("user_name"),
        highest_income=highest_income,
        highest_expense=highest_expense,
        average_expense=average_expense,
        most_used_category=most_used_category,
        budget_data=budget_data,
        monthly_income=monthly_income,
        monthly_expense=monthly_expense,
        monthly_savings=monthly_savings,
        top_expenses=top_expenses,
        category_percentages=category_percentages,
        recurring_transactions=recurring_transactions,
        bill_reminders=bill_reminders,
        financial_insights=financial_insights,
        ai_predictions=ai_predictions
    )


@app.route("/transactions", methods=["GET", "POST"])
@login_required
def transactions():
    init_db()
    user_id = session.get("user_id")

    if request.method == "POST":
        title = request.form["title"]
        amount = request.form["amount"]
        category = request.form["category"]
        trans_type = request.form["type"]
        date = request.form["date"]

        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO transactions (user_id, title, amount, category, type, date)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (user_id, title, amount, category, trans_type, date))
        conn.commit()
        conn.close()

        flash("Transaction added successfully.")
        return redirect("/transactions")

    search = request.args.get("search", "")
    category = request.args.get("category", "")
    trans_type = request.args.get("type", "")
    from_date = request.args.get("from_date", "")
    to_date = request.args.get("to_date", "")
    sort = request.args.get("sort", "newest")

    query = "SELECT * FROM transactions WHERE user_id=?"
    params = [user_id]

    if search:
        query += " AND title LIKE ?"
        params.append(f"%{search}%")

    if category:
        query += " AND category = ?"
        params.append(category)

    if trans_type:
        query += " AND type = ?"
        params.append(trans_type)

    if from_date:
        query += " AND date >= ?"
        params.append(from_date)

    if to_date:
        query += " AND date <= ?"
        params.append(to_date)

    if sort == "oldest":
        query += " ORDER BY date ASC"
    elif sort == "highest":
        query += " ORDER BY amount DESC"
    elif sort == "lowest":
        query += " ORDER BY amount ASC"
    else:
        query += " ORDER BY date DESC"

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(query, params)
    all_transactions = cursor.fetchall()
    conn.close()

    return render_template(
        "transactions.html",
        transactions=all_transactions,
        search=search,
        selected_category=category,
        selected_type=trans_type,
        from_date=from_date,
        to_date=to_date,
        selected_sort=sort
    )


@app.route("/analytics")
@login_required
def analytics():
    income, expense, balance, total_transactions = get_summary()
    chart_data = get_chart_data()

    return render_template(
        "analytics.html",
        income=income,
        expense=expense,
        balance=balance,
        total_transactions=total_transactions,
        chart_data=chart_data
    )


@app.route("/reports")
@login_required
def reports():
    income, expense, balance, total_transactions = get_summary()

    return render_template(
        "reports.html",
        income=income,
        expense=expense,
        balance=balance,
        total_transactions=total_transactions
    )


@app.route("/export/csv")
@login_required
def export_csv():
    user_id = session.get("user_id")
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT title, amount, category, type, date
        FROM transactions
        WHERE user_id=?
        ORDER BY date DESC
    """, (user_id,))

    transactions = cursor.fetchall()
    conn.close()

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["Title", "Amount", "Category", "Type", "Date"])

    for transaction in transactions:
        writer.writerow(transaction)

    flash("CSV report downloaded.")
    response = Response(output.getvalue(), mimetype="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=expense_report.csv"
    return response


@app.route("/export/excel")
@login_required
def export_excel():
    user_id = session.get("user_id")
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT title, amount, category, type, date
        FROM transactions
        WHERE user_id=?
        ORDER BY date DESC
    """, (user_id,))

    transactions = cursor.fetchall()
    conn.close()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Expense Report"

    sheet.append(["Title", "Amount", "Category", "Type", "Date"])

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")

    for transaction in transactions:
        sheet.append(transaction)

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        sheet.column_dimensions[column_letter].width = max_length + 5

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    flash("Excel report downloaded.")
    response = Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response.headers["Content-Disposition"] = "attachment; filename=expense_report.xlsx"
    return response


@app.route("/export/pdf")
@login_required
def export_pdf():
    user_id = session.get("user_id")
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT title, amount, category, type, date
        FROM transactions
        WHERE user_id=?
        ORDER BY date DESC
    """, (user_id,))

    transactions = cursor.fetchall()
    conn.close()

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph("Personal Expense Tracker Report", styles["Title"]))
    elements.append(Spacer(1, 20))

    income, expense, balance, total_transactions = get_summary()

    elements.append(
        Paragraph(
            f"Total Income: Rs. {income}<br/>"
            f"Total Expense: Rs. {expense}<br/>"
            f"Balance: Rs. {balance}<br/>"
            f"Total Transactions: {total_transactions}",
            styles["Normal"]
        )
    )

    elements.append(Spacer(1, 20))

    data = [["Title", "Amount", "Category", "Type", "Date"]]

    for t in transactions:
        data.append([t[0], f"Rs. {t[1]}", t[2], t[3], t[4]])

    table = Table(data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 1, colors.grey),
        ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
    ]))

    elements.append(table)
    doc.build(elements)
    output.seek(0)

    flash("PDF report downloaded.")
    return Response(
        output.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": "attachment; filename=expense_report.pdf"}
    )


@app.route("/delete/<int:id>")
@login_required
def delete_transaction(id):
    user_id = session.get("user_id")
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM transactions WHERE id=? AND user_id=?", (id, user_id))
    conn.commit()
    conn.close()

    flash("Transaction deleted successfully.")
    return redirect("/transactions")


@app.route("/edit/<int:id>", methods=["GET", "POST"])
@login_required
def edit_transaction(id):
    user_id = session.get("user_id")
    conn = get_connection()
    cursor = conn.cursor()

    if request.method == "POST":
        title = request.form["title"]
        amount = request.form["amount"]
        category = request.form["category"]
        trans_type = request.form["type"]
        date = request.form["date"]

        cursor.execute("""
            UPDATE transactions
            SET title=?, amount=?, category=?, type=?, date=?
            WHERE id=? AND user_id=?
        """, (title, amount, category, trans_type, date, id, user_id))

        conn.commit()
        conn.close()

        flash("Transaction updated successfully.")
        return redirect("/transactions")

    cursor.execute("SELECT * FROM transactions WHERE id=? AND user_id=?", (id, user_id))
    transaction = cursor.fetchone()
    conn.close()

    if transaction is None:
        flash("Transaction not found.")
        return redirect("/transactions")

    return render_template("edit.html", transaction=transaction)

@app.route("/calendar")
@login_required
def calendar_view():
    calendar_items = get_calendar_data()

    return render_template(
        "calendar.html",
        calendar_items=calendar_items
    )
if __name__ == "__main__":
    init_db()
    app.run(debug=True)